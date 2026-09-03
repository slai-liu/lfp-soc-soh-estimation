"""Load the CALCE A123 drive-cycle archives used by the SOC reproduction."""
from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd


@dataclass
class A123Segment:
    profile: str
    ambient_temperature_c: int
    time_s: np.ndarray
    voltage_v: np.ndarray
    current_a: np.ndarray
    measured_temperature_c: np.ndarray
    soc: np.ndarray
    usable_capacity_ah: float
    archive: str
    workbook: str
    source_rows: int

    @property
    def sample_interval_s(self) -> float:
        return float(np.median(np.diff(self.time_s)))

    def audit_row(self) -> dict[str, object]:
        return {
            "profile": self.profile,
            "ambient_temperature_c": self.ambient_temperature_c,
            "samples": len(self.time_s),
            "duration_s": float(self.time_s[-1]),
            "median_sample_interval_s": self.sample_interval_s,
            "voltage_min_v": float(self.voltage_v.min()),
            "voltage_max_v": float(self.voltage_v.max()),
            "current_min_a": float(self.current_a.min()),
            "current_max_a": float(self.current_a.max()),
            "usable_capacity_ah": self.usable_capacity_ah,
            "soc_start": float(self.soc[0]),
            "soc_end": float(self.soc[-1]),
            "archive": self.archive,
            "workbook": self.workbook,
            "source_rows": self.source_rows,
        }


def _header(value: object) -> str:
    return str(value).strip().replace(" ", "").lower()


def archive_name(temperature_c: int) -> str:
    return f"A123_DST-US06-FUDS-{temperature_c}.zip"


def _select_workbook(zipped: zipfile.ZipFile, cell: str, temperature_c: int) -> str:
    candidates = [
        name for name in zipped.namelist()
        if name.lower().endswith(".xlsx")
        and not Path(name).name.startswith("~$")
        and cell.lower() in Path(name).name.lower()
    ]
    if temperature_c == 20:
        revised = [name for name in candidates if "newprofile" in name.lower()]
        if revised:
            candidates = revised
    if not candidates:
        raise FileNotFoundError(f"No workbook for {cell} at {temperature_c} °C")
    return sorted(candidates)[0]


def _read_workbook(payload: bytes) -> tuple[np.ndarray, ...]:
    # Keep the spreadsheet dependency on the real-data path so the synthetic
    # self-test can run before users download or install data-specific extras.
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheets = [sheet for sheet in workbook.worksheets if sheet.title.lower().startswith("channel_")]
    if not sheets:
        workbook.close()
        raise ValueError("No Channel_* worksheet found")
    sheet = max(sheets, key=lambda item: item.max_row)
    rows = sheet.iter_rows(values_only=True)
    columns = {name: index for index, name in enumerate(map(_header, next(rows)))}
    required = ["test_time(s)", "step_index", "current(a)", "voltage(v)"]
    missing = [name for name in required if name not in columns]
    if missing:
        workbook.close()
        raise KeyError(f"Missing columns: {missing}")
    temperature_columns = [i for name, i in columns.items() if name.startswith("temperature(c)")]
    values: list[list[float]] = [[], [], [], [], []]
    for row in rows:
        try:
            values[0].append(float(row[columns["test_time(s)"]]))
            values[1].append(float(row[columns["step_index"]]))
            values[2].append(float(row[columns["current(a)"]]))
            values[3].append(float(row[columns["voltage(v)"]]))
            values[4].append(float(row[temperature_columns[0]]) if temperature_columns else np.nan)
        except (TypeError, ValueError, IndexError):
            continue
    workbook.close()
    return tuple(np.asarray(item, dtype=np.float64) for item in values)


def _segment(arrays: tuple[np.ndarray, ...], *, profile: str,
             step_range: tuple[int, int], temperature_c: int,
             archive: str, workbook: str) -> A123Segment:
    time, step, current, voltage, measured_temperature = arrays
    starts = np.flatnonzero(step == step_range[0])
    if not len(starts):
        raise ValueError(f"Step {step_range[0]} not found in {workbook}")
    start = int(starts[0])
    stops = np.flatnonzero((np.arange(len(step)) > start) & (step >= step_range[1]))
    stop = int(stops[0]) if len(stops) else len(step)
    source_rows = stop - start
    time, current, voltage, measured_temperature = (
        item[start:stop] for item in (time, current, voltage, measured_temperature)
    )
    finite = np.isfinite(time) & np.isfinite(current) & np.isfinite(voltage)
    time, current, voltage, measured_temperature = (
        item[finite] for item in (time, current, voltage, measured_temperature)
    )
    measured_temperature = np.where(np.isfinite(measured_temperature), measured_temperature, temperature_c)
    order = np.argsort(time, kind="stable")
    time, current, voltage, measured_temperature = (
        item[order] for item in (time, current, voltage, measured_temperature)
    )
    increasing = np.concatenate(([True], np.diff(time) > 0))
    time, current, voltage, measured_temperature = (
        item[increasing] for item in (time, current, voltage, measured_temperature)
    )
    time = time - time[0]
    dt = np.diff(time, prepend=time[0])
    discharged_ah = -np.cumsum(current * dt) / 3600.0
    discharged_ah -= discharged_ah[0]
    capacity = float(np.max(discharged_ah))
    if capacity <= 0:
        raise ValueError(f"Non-positive usable capacity for {profile} at {temperature_c} °C")
    soc = np.clip(1.0 - discharged_ah / capacity, 0.0, 1.0)
    return A123Segment(profile, temperature_c, time.astype(np.float64), voltage.astype(np.float32),
                       current.astype(np.float32), measured_temperature.astype(np.float32),
                       soc.astype(np.float32), capacity, archive, workbook, source_rows)


def load_segments(data_dir: Path, config: dict[str, object]) -> list[A123Segment]:
    segments: list[A123Segment] = []
    for temperature in map(int, config["temperatures_c"]):
        archive_path = data_dir / archive_name(temperature)
        if not archive_path.is_file():
            raise FileNotFoundError(archive_path)
        with zipfile.ZipFile(archive_path) as zipped:
            workbook = _select_workbook(zipped, str(config["cell"]), temperature)
            arrays = _read_workbook(zipped.read(workbook))
            for profile, bounds in dict(config["profile_steps"]).items():
                segments.append(_segment(arrays, profile=profile,
                                         step_range=tuple(map(int, bounds)),
                                         temperature_c=temperature,
                                         archive=archive_path.name,
                                         workbook=Path(workbook).name))
    return segments


def validate_segments(segments: list[A123Segment], config: dict[str, object]) -> None:
    expected = {(p, int(t)) for p in config["profile_steps"] for t in config["temperatures_c"]}
    actual = {(s.profile, s.ambient_temperature_c) for s in segments}
    if actual != expected:
        raise ValueError(f"Segment mismatch: missing={expected-actual}, extra={actual-expected}")
    for segment in segments:
        arrays = [segment.time_s, segment.voltage_v, segment.current_a,
                  segment.measured_temperature_c, segment.soc]
        if any(not np.all(np.isfinite(item)) for item in arrays):
            raise ValueError(f"Non-finite values in {segment.profile} at {segment.ambient_temperature_c} °C")
        if len(segment.time_s) < 1000 or np.any(np.diff(segment.time_s) <= 0):
            raise ValueError("Invalid time axis")


def write_data_audit(segments: list[A123Segment], output_path: Path) -> pd.DataFrame:
    frame = pd.DataFrame([item.audit_row() for item in segments]).sort_values(
        ["profile", "ambient_temperature_c"])
    frame.to_csv(output_path, index=False)
    return frame


def select_profile(segments: list[A123Segment], profile: str) -> list[A123Segment]:
    return sorted((item for item in segments if item.profile == profile),
                  key=lambda item: item.ambient_temperature_c)
